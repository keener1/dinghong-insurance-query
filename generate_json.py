# -*- coding: utf-8 -*-
import sys, os, codecs, json, glob
if os.name == 'nt':
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

from openpyxl import load_workbook
from datetime import datetime

# 自动查找Downloads目录下最新的推广费Excel
def find_latest_excel():
    patterns = [
        os.path.join(r'C:\Users\15936\Downloads', '*推广费*.xlsx'),
        os.path.join(r'C:\Users\15936\Downloads', '*推广*.xlsx'),
    ]
    all_files = []
    for p in patterns:
        all_files.extend(glob.glob(p))
    if not all_files:
        return None
    return max(all_files, key=os.path.getmtime)

def pf(fee):
    if fee is None or fee == '/':
        return 0
    try:
        return float(str(fee).replace('%', ''))
    except:
        return 0

def parse_excel(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['表1-费用明细']
    headers = [cell.value for cell in ws[2]]

    # 动态检测所有"第X年总计"列
    year_cols = {}
    for i, h in enumerate(headers):
        if h and '年总计' in str(h) and '第' in str(h):
            year_num = str(h).replace('第', '').replace('年总计', '')
            year_cols[year_num] = i + 1

    years_list = sorted(year_cols.keys(), key=lambda x: int(x))
    print(f"  检测到年份列: {years_list}")

    products = {}
    for row_idx in range(3, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        if not row[0]:
            continue
        
        prod_id = row[1]
        ins_type = str(row[2]).strip() if row[2] else ""
        prod_name = str(row[3]).strip() if row[3] else ""
        fee_rate = row[10]
        fee_combo = str(row[7]).strip() if row[7] else ""
        attr = str(row[0]).strip()
        
        key = f"{prod_id}_{prod_name}" if prod_id else prod_name
        if key not in products:
            products[key] = {
                'id': prod_id,
                'name': prod_name,
                'type': ins_type,
                'attr': attr,
                'total_fee': fee_rate,
                'years': years_list,
                'variants': []
            }
        
        yearly_totals = []
        for y in years_list:
            col_idx = year_cols[y]
            val = row[col_idx - 1] if col_idx <= len(row) else None
            if val is None or val == '/':
                yearly_totals.append('0')
            else:
                yearly_totals.append(str(val))
        
        current_max = pf(products[key]['total_fee'])
        if pf(fee_rate) > current_max:
            products[key]['total_fee'] = fee_rate
        
        products[key]['variants'].append({
            'fee_combo': fee_combo,
            'total_fee': str(fee_rate) if fee_rate else '0',
            'yearly_totals': yearly_totals,
        })
    
    return products, years_list

def compare_with_previous(current_products, old_data=None):
    """与上一次数据对比，生成变更报告"""
    if old_data is None:
        return None
    
    prev_products = old_data.get('products', [])
    
    # 构建上一次产品字典（以id+name为key）
    prev_dict = {}
    for p in prev_products:
        key = f"{p.get('id','')}_{p.get('name','')}"
        prev_dict[key] = p
    
    # 当前产品字典
    curr_dict = {}
    for key, p in current_products.items():
        curr_dict[key] = p
    
    prev_keys = set(prev_dict.keys())
    curr_keys = set(curr_dict.keys())
    
    # 下架产品（之前有，现在没有）
    delisted = []
    for key in prev_keys - curr_keys:
        p = prev_dict[key]
        delisted.append({
            'name': p.get('name', ''),
            'type': p.get('type', ''),
        })
    
    # 上架产品（现在有，之前没有）
    new_products = []
    for key in curr_keys - prev_keys:
        p = curr_dict[key]
        new_products.append({
            'name': p.get('name', ''),
            'type': p.get('type', ''),
        })
    
    # 费用变动产品（两边都有但费用不同）
    fee_changes = []
    for key in prev_keys & curr_keys:
        prev_p = prev_dict[key]
        curr_p = curr_dict[key]
        
        prev_max_fee = pf(prev_p.get('total_fee', 0))
        curr_max_fee = pf(curr_p.get('total_fee', 0))
        
        if prev_max_fee != curr_max_fee:
            fee_changes.append({
                'name': curr_p.get('name', ''),
                'type': curr_p.get('type', ''),
                'old_fee': str(prev_p.get('total_fee', '')),
                'new_fee': str(curr_p.get('total_fee', '')),
                'old_fee_num': prev_max_fee,
                'new_fee_num': curr_max_fee,
            })
    
    # 按变动幅度排序（变化大的在前）
    fee_changes.sort(key=lambda x: abs(x['new_fee_num'] - x['old_fee_num']), reverse=True)
    
    has_changes = len(delisted) > 0 or len(new_products) > 0 or len(fee_changes) > 0
    
    return {
        'date': datetime.now().strftime('%Y-%m-%d'),
        'time': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'has_changes': has_changes,
        'delisted_count': len(delisted),
        'new_count': len(new_products),
        'fee_change_count': len(fee_changes),
        'delisted': delisted,
        'new_products': new_products,
        'fee_changes': fee_changes,
    }

if __name__ == '__main__':
    excel_path = find_latest_excel()
    if not excel_path:
        print("ERROR: 未找到推广费Excel文件！")
        sys.exit(1)
    print(f"使用Excel文件: {os.path.basename(excel_path)}")
    
    # 先读旧数据（在覆盖之前）
    json_path = r'C:\Users\15936\WorkBuddy\20260326151224\niubao_products.json'
    old_data = None
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                old_data = json.load(f)
        except:
            pass
    
    # 解析新Excel
    products, years_list = parse_excel(excel_path)
    
    # 变更对比（用旧数据对比新数据）
    changes = compare_with_previous(products, old_data)
    
    sorted_products = sorted(products.values(), key=lambda x: pf(x['total_fee']), reverse=True)
    
    output = {
        'updateTime': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'totalCount': len(sorted_products),
        'products': sorted_products,
    }
    
    json_path = r'C:\Users\15936\WorkBuddy\20260326151224\niubao_products.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"Done! {len(sorted_products)} products saved")
    
    # 保存变更报告
    if changes:
        changes_path = r'C:\Users\15936\WorkBuddy\20260326151224\niubao_changes.json'
        with open(changes_path, 'w', encoding='utf-8') as f:
            json.dump(changes, f, ensure_ascii=False, indent=2)
        print(f"变更报告: 下架{changes['delisted_count']} 上架{changes['new_count']} 费用变动{changes['fee_change_count']}")
        if not changes['has_changes']:
            print("今日无变动")
    else:
        print("首次运行或无历史数据，跳过变更对比")
