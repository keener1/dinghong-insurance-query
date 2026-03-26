# -*- coding: utf-8 -*-
import sys, os, codecs, json, glob
if os.name == 'nt':
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

from openpyxl import load_workbook

def find_latest_excel():
    patterns = [
        os.path.join(r'C:\Users\15936\Downloads', '*推广费*.xlsx'),
        os.path.join(r'C:\Users\15936\Downloads', '*推广*.xlsx'),
    ]
    all_files = []
    for p in patterns:
        all_files.extend(glob.glob(p))
    if not all_files:
        return None
    return max(all_files, key=os.path.getmtime)

def pf(fee):
    if fee is None or fee == '/':
        return 0
    try:
        return float(str(fee).replace('%', ''))
    except:
        return 0

excel_path = find_latest_excel()
if not excel_path:
    print("ERROR: 未找到推广费Excel文件！")
    sys.exit(1)
print(f"使用Excel文件: {os.path.basename(excel_path)}")
wb = load_workbook(excel_path, data_only=True)

all_data = {}

# ===== Sheet 1: 费用明细 =====
ws = wb['表1-费用明细']
headers = [cell.value for cell in ws[2]]
products = {}
types_count = {}

for row_idx in range(3, ws.max_row + 1):
    row = [cell.value for cell in ws[row_idx]]
    if not row[0]:
        continue
    
    prod_id = row[1]
    ins_type = str(row[2]) if row[2] else ""
    prod_name = str(row[3]) if row[3] else ""
    fee_rate = row[10]   # 总费用
    y1_total = row[11]   # 第1年总计
    y1_base = row[12]    # 第1年基础
    y1_act = row[13]     # 第1年活动
    y2_total = row[14]   # 第2年总计
    y3_total = row[19]   # 第3年总计
    y4_total = row[24]   # 第4年总计
    y5_total = row[25]   # 第5年总计
    responsibility = row[6]  # 责任
    fee_combo = row[7]       # 费率组合
    settle_req = row[8]      # 基本结算要求
    hesitation = row[9]      # 犹豫期
    attr = str(row[0])       # 商品属性
    
    key = f"{prod_id}_{prod_name}" if prod_id else prod_name
    if key not in products:
        products[key] = {
            'id': prod_id,
            'name': prod_name,
            'type': ins_type,
            'attr': attr,
            'total_fee': fee_rate,
            'max_total_fee': pf(fee_rate),
            'y1_total': y1_total,
            'y1_base': y1_base,
            'y1_activity': y1_act,
            'y2_total': y2_total,
            'y3_total': y3_total,
            'y4_total': y4_total,
            'y5_total': y5_total,
            'responsibility': responsibility,
            'fee_combo': fee_combo,
            'settle_requirement': settle_req,
            'hesitation_period': hesitation,
            'variants': [{
                'fee_combo': fee_combo,
                'total_fee': fee_rate,
                'y1_total': y1_total,
                'responsibility': responsibility,
            }]
        }
    else:
        if pf(fee_rate) > products[key]['max_total_fee']:
            products[key]['max_total_fee'] = pf(fee_rate)
            products[key]['total_fee'] = fee_rate
        products[key]['variants'].append({
            'fee_combo': fee_combo,
            'total_fee': fee_rate,
            'y1_total': y1_total,
            'responsibility': responsibility,
        })
    
    if ins_type and ins_type != 'None':
        types_count[ins_type] = types_count.get(ins_type, 0) + 1

# 按险种分组
by_type = {}
for p in products.values():
    t = p['type']
    if t not in by_type:
        by_type[t] = []
    by_type[t].append(p)

# 按总费用排序
all_sorted = sorted(products.values(), key=lambda x: x['max_total_fee'], reverse=True)

all_data['sheet1'] = {
    'total_products': len(products),
    'total_rows': ws.max_row - 2,
    'types': types_count,
    'top30': [{
        'id': p['id'], 'name': p['name'], 'type': p['type'],
        'total_fee': p['total_fee'], 'max_fee': p['max_total_fee'],
        'y1_total': str(p['y1_total']), 'y1_base': str(p['y1_base']),
        'y1_activity': str(p['y1_activity']), 'y2_total': str(p['y2_total']),
        'variants_count': len(p['variants']),
        'settle_requirement': str(p['settle_requirement'])[:80] if p['settle_requirement'] else '',
        'hesitation_period': str(p['hesitation_period']) if p['hesitation_period'] else '',
    } for p in all_sorted[:30]],
    'by_type_summary': {t: {
        'count': len(ps),
        'top5': [{'name': p['name'], 'total_fee': p['total_fee'], 'max_fee': p['max_total_fee']} for p in sorted(ps, key=lambda x: x['max_total_fee'], reverse=True)[:5]]
    } for t, ps in by_type.items()},
    'all_products': [{
        'id': p['id'], 'name': p['name'], 'type': p['type'],
        'attr': p['attr'], 'total_fee': p['total_fee'], 'max_fee': p['max_total_fee'],
        'y1_total': str(p['y1_total']), 'y1_base': str(p['y1_base']),
        'y1_activity': str(p['y1_activity']), 'y2_total': str(p['y2_total']),
        'y3_total': str(p['y3_total']), 'y4_total': str(p['y4_total']),
        'y5_total': str(p['y5_total']),
        'settle_requirement': str(p['settle_requirement'])[:100] if p['settle_requirement'] else '',
        'hesitation_period': str(p['hesitation_period']) if p['hesitation_period'] else '',
        'variants': p['variants'][:3],
    } for p in all_sorted],
}

# ===== Sheet 2: 限时补贴活动 =====
ws2 = wb['表2-限时补贴活动']
headers2 = [cell.value for cell in ws2[2]]
activities2 = []
for row_idx in range(3, ws2.max_row + 1):
    row = [cell.value for cell in ws2[row_idx]]
    if not row[0]:
        continue
    activities2.append({
        'product': str(row[2]) if row[2] else '',
        'activity_name': str(row[3]) if row[3] else '',
        'details': [str(c) if c else '' for c in row[:8]],
    })

all_data['sheet2_limited_time_activities'] = {
    'total': len(activities2),
    'headers': headers2[:8],
    'activities': activities2[:50],
}

# ===== Sheet 3: 阶梯奖励活动 =====
ws3 = wb['表3-阶梯奖励活动']
headers3 = [cell.value for cell in ws3[2]]
activities3 = []
for row_idx in range(3, ws3.max_row + 1):
    row = [cell.value for cell in ws3[row_idx]]
    if not row[0]:
        continue
    activities3.append({
        'product': str(row[2]) if row[2] else '',
        'activity_name': str(row[3]) if row[3] else '',
        'details': [str(c) if c else '' for c in row[:10]],
    })

all_data['sheet3_tiered_rewards'] = {
    'total': len(activities3),
    'headers': headers3[:10],
    'activities': activities3[:50],
}

# ===== Sheet 4: 继续率考核 =====
ws4 = wb['表4-继续率考核']
headers4 = [cell.value for cell in ws4[2]]
cont_items = []
for row_idx in range(3, ws4.max_row + 1):
    row = [cell.value for cell in ws4[row_idx]]
    if not row[0]:
        continue
    cont_items.append({
        'product': str(row[2]) if row[2] else '',
        'details': [str(c) if c else '' for c in row[:12]],
    })

all_data['sheet4_continuation_rate'] = {
    'total': len(cont_items),
    'headers': headers4[:12],
    'items': cont_items[:50],
}

# Save
output_path = r'C:\Users\15936\WorkBuddy\20260326151224\niubao_data_analysis.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"Data saved to: {output_path}")
print(f"Total unique products: {len(products)}")
print(f"Insurance types: {types_count}")
print("Done!")
