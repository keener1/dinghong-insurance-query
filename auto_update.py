# -*- coding: utf-8 -*-
"""
鼎鸿保险经纪费用查询 - 全自动更新脚本
流程：浏览器登录牛保100 → 导出商品推广费 → 等待导出 → 下载Excel → 生成JSON → 复制index.html → git推送
"""
import sys, os, codecs, json, glob, time, shutil
if os.name == 'nt':
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

from openpyxl import load_workbook
from datetime import datetime

WORKSPACE = r'C:\Users\15936\WorkBuddy\20260326151224'
DOWNLOADS = r'C:\Users\15936\Downloads'
MCPORTER = r'C:\Users\15936\.agents\skills\autoglm-browser-agent\dependency\mcporter.exe'
RELAY = r'C:\Users\15936\.agents\skills\autoglm-browser-agent\dist\relay.exe'

def step1_launch_relay():
    """启动relay服务"""
    print("[1/7] 启动relay服务...")
    # 检查relay是否已在运行
    import subprocess
    try:
        result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq relay.exe'], 
                              capture_output=True, text=True)
        if 'relay.exe' in result.stdout:
            print("  relay已在运行")
            return True
    except:
        pass
    
    try:
        subprocess.Popen([RELAY, '62030', '62031'], 
                        creationflags=subprocess.CREATE_NO_WINDOW)
        print("  relay启动成功")
        time.sleep(2)
        return True
    except Exception as e:
        print(f"  relay启动失败: {e}")
        return False

def step2_export_data():
    """浏览器自动化：登录+导出+下载"""
    print("[2/7] 浏览器自动化：登录牛保100并导出数据...")
    
    now = datetime.now()
    year = now.year
    month = now.month
    
    task = (
        f"登录牛保100保险平台，账号13376413472密码Dhtx123456@，"
        f"登录成功后进入商品库页面，"
        f"在查询按钮右边找到'导出商品推广费'按钮并点击，"
        f"在弹出窗口中选择{year}年{month}月的时间范围并导出，"
        f"导出后等待，然后去数据分析栏的导表中心，"
        f"找到最新的导出记录（检查时间是否是今天的），"
        f"确认状态为'已完成'后，点击下载该Excel文件"
    )
    
    import subprocess
    try:
        result = subprocess.run(
            [MCPORTER, 'call', 'autoglm-browser-agent.browser_subagent',
             f'task={task}',
             'start_url=https://www.niubao100.com/index/main#/login/signIn?redirect=%2Fhome',
             '--timeout', '600000'],  # 10分钟超时
            capture_output=True, text=True, timeout=600
        )
        print(f"  浏览器自动化执行完成")
        if result.stdout:
            print(f"  输出: {result.stdout[-500:]}")
        return True
    except subprocess.TimeoutExpired:
        print("  浏览器自动化超时")
        return False
    except Exception as e:
        print(f"  浏览器自动化执行失败: {e}")
        return False

def step3_find_excel():
    """找到最新下载的推广费Excel"""
    print("[3/7] 查找最新下载的Excel文件...")
    
    # 查找所有推广费Excel
    files = glob.glob(os.path.join(DOWNLOADS, '*推广费*.xlsx'))
    if not files:
        files = glob.glob(os.path.join(DOWNLOADS, '*推广*.xlsx'))
    
    if not files:
        print("  未找到推广费Excel文件！")
        return None
    
    # 取最新的
    latest = max(files, key=os.path.getmtime)
    mtime = datetime.fromtimestamp(os.path.getmtime(latest))
    print(f"  找到: {os.path.basename(latest)} (修改时间: {mtime})")
    return latest

def step4_generate_json(excel_path):
    """从Excel生成网页所需的JSON"""
    print("[4/7] 解析Excel并生成JSON...")
    
    def pf(fee):
        if fee is None or fee == '/':
            return 0
        try:
            return float(str(fee).replace('%', ''))
        except:
            return 0
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['表1-费用明细']
    headers = [cell.value for cell in ws[2]]
    
    # 动态检测所有"第X年总计"列
    year_cols = {}
    for i, h in enumerate(headers):
        if h and '年总计' in str(h) and '第' in str(h):
            year_num = str(h).replace('第', '').replace('年总计', '')
            year_cols[year_num] = i + 1
    
    years_list = sorted(year_cols.keys(), key=lambda x: int(x))
    print(f"  检测到年份列: {years_list}")
    
    products = {}
    for row_idx in range(3, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        if not row[0]:
            continue
        
        prod_id = row[1]
        ins_type = str(row[2]).strip() if row[2] else ""
        prod_name = str(row[3]).strip() if row[3] else ""
        fee_rate = row[10]
        fee_combo = str(row[7]).strip() if row[7] else ""
        attr = str(row[0]).strip()
        
        key = f"{prod_id}_{prod_name}" if prod_id else prod_name
        if key not in products:
            products[key] = {
                'id': prod_id,
                'name': prod_name,
                'type': ins_type,
                'attr': attr,
                'total_fee': fee_rate,
                'years': years_list,
                'variants': []
            }
        
        yearly_totals = []
        for y in years_list:
            col_idx = year_cols[y]
            val = row[col_idx - 1] if col_idx <= len(row) else None
            if val is None or val == '/':
                yearly_totals.append('0')
            else:
                yearly_totals.append(str(val))
        
        current_max = pf(products[key]['total_fee'])
        if pf(fee_rate) > current_max:
            products[key]['total_fee'] = fee_rate
        
        products[key]['variants'].append({
            'fee_combo': fee_combo,
            'total_fee': str(fee_rate) if fee_rate else '0',
            'yearly_totals': yearly_totals,
        })
    
    sorted_products = sorted(products.values(), key=lambda x: pf(x['total_fee']), reverse=True)
    
    output = {
        'updateTime': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'totalCount': len(sorted_products),
        'products': sorted_products,
    }
    
    json_path = os.path.join(WORKSPACE, 'niubao_products.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"  生成JSON完成: {len(sorted_products)}个产品")
    return json_path

def step5_copy_index():
    """复制niubao_query.html为index.html"""
    print("[5/7] 更新index.html...")
    src = os.path.join(WORKSPACE, 'niubao_query.html')
    dst = os.path.join(WORKSPACE, 'index.html')
    shutil.copy2(src, dst)
    print("  index.html已更新")
    return True

def step6_git_push():
    """Git提交并推送"""
    print("[6/7] Git提交并推送到GitHub...")
    import subprocess
    
    try:
        subprocess.run(['git', 'add', 'niubao_products.json', 'index.html'], 
                      cwd=WORKSPACE, capture_output=True, text=True)
        result = subprocess.run(
            ['git', 'commit', '-m', f'自动更新: {datetime.now().strftime("%Y-%m-%d %H:%M")}'],
            cwd=WORKSPACE, capture_output=True, text=True
        )
        if 'nothing to commit' in result.stdout:
            print("  没有变更需要提交")
            return True
        
        result = subprocess.run(['git', 'push'], cwd=WORKSPACE, 
                              capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            print("  推送成功！")
            return True
        else:
            print(f"  推送失败: {result.stderr}")
            return False
    except Exception as e:
        print(f"  Git操作失败: {e}")
        return False

def step7_report():
    """输出最终报告"""
    print("\n[7/7] 更新完成！")
    print(f"  网页地址: https://keener1.github.io/dinghong-insurance-query/")
    print(f"  更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

if __name__ == '__main__':
    print(f"===== 鼎鸿保险经纪费用查询 - 自动更新开始 =====")
    print(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    step1_launch_relay()
    step2_export_data()
    
    excel_path = step3_find_excel()
    if not excel_path:
        print("错误：未找到Excel文件，退出")
        sys.exit(1)
    
    step4_generate_json(excel_path)
    step5_copy_index()
    step6_git_push()
    step7_report()
    
    print("\n===== 全流程完成 =====")
